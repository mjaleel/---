import streamlit as st
import pandas as pd
import math
import os
import json
import csv
import io
import zipfile
from datetime import datetime
from itertools import groupby
from num2words import num2words
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ====================================================================
# الثوابت والبيانات الثابتة
# ====================================================================

PAYER_NAME = "مديرية تربية البصرة"
PAYER_ACCOUNT = "IQ26RAFB002100366585001"
CURRENCY = "IQD"
DETAILS_OF_CHARGES = "SLEV"
REMITTANCE_INFO_TEMPLATE = "SALARY {} {}"
MAX_ROWS_PER_FILE = 4000
MAX_AMOUNT_PER_FILE = 4_000_000_000

BANK_BICS = {
    'RAFB': 'RAFBIQB1098', 'RDBA': 'RDBAIQB1046', 'AIBI': 'AIBIIQBA991',
    'IDBQ': 'IDBQIQBA004', 'AINI': 'AINIIQBA015', 'NBIQ': 'NBIQIQBA830'
}

ARABIC_BANK_NAME_MAP = {
    'RAFB': 'الرافدين', 'RDBA': 'الرشيد', 'AIBI': 'آشور',
    'IDBQ': 'التنمية', 'AINI': 'الطيف', 'NBIQ': 'الأهلي'
}

BANKS_WITH_DYNAMIC_BRANCHES = ['AINI', 'NBIQ', 'AIBI']
BANK_KEYS_FOR_FILTERING = list(BANK_BICS.keys())

ARABIC_MONTHS = {
    1: "كانون الثاني", 2: "شباط", 3: "آذار", 4: "نيسان", 5: "أيار", 6: "حزيران",
    7: "تموز", 8: "آب", 9: "أيلول", 10: "تشرين الأول", 11: "تشرين الثاني", 12: "كانون الأول"
}

FINAL_EXCEL_COLS = [
    'Reference', 'Value Date', 'Payer Name', 'Payer Account', 'Amount',
    'Currency', 'Receiver BIC', 'Beneficiary Name', 'Beneficiary Account',
    'Remittance Information', 'Details of Charges'
]

ARABIC_DIGITS = str.maketrans('0123456789', '٠١٢٣٤٥٦٧٨٩')

DEFAULT_BRANCHES = {
    "NBIQ": ["NBIQIQBA830", "NBIQIQBA863"],
    "AIBI": ["AIBIIQBA991", "AIBIIQBA988"]
}

# ====================================================================
# دوال مساعدة
# ====================================================================

def convert_amount_to_arabic(amount):
    try:
        val = float(amount)
        text = num2words(int(val), lang='ar')
        text = text.replace("ألفاً", "ألفًا").replace("مليوناً", "مليونًا").replace("ملياراً", "مليارًا")
        text = text.replace(" و ", " و").strip()
        return text + " دينار لا غير"
    except:
        return "صفر دينار"


def format_number(value, use_arabic_digits=False):
    try:
        formatted = f"{int(value):,}"
        if use_arabic_digits:
            formatted = formatted.translate(ARABIC_DIGITS)
            formatted = formatted.replace(',', '،')
        return formatted
    except:
        return str(value)


def guess_column(keywords, columns):
    for kw in keywords:
        for col in columns:
            if kw.lower() in str(col).lower():
                return col
    return columns[0] if columns else ""


def apply_financial_rounding(df, column_name):
    df = df.copy().reset_index(drop=False)
    original_total = round(df[column_name].sum(), 2)
    df['_floor_val'] = df[column_name].apply(math.floor)
    df['_decimal_part'] = df[column_name] - df['_floor_val']
    total_to_distribute = int(round(original_total - df['_floor_val'].sum()))
    df_sorted = df.sort_values(by='_decimal_part', ascending=False).copy()
    df_sorted['_final_amount'] = df_sorted['_floor_val']
    if total_to_distribute > 0:
        top_positions = df_sorted.index[:total_to_distribute]
        df_sorted.loc[top_positions, '_final_amount'] += 1
    df_sorted = df_sorted.sort_values(by='index').set_index('index')
    df_sorted.index.name = None
    df_sorted[column_name] = df_sorted['_final_amount'].astype(int)
    return df_sorted.drop(columns=['_floor_val', '_decimal_part', '_final_amount'])


def get_receiver_bic_dynamic(row, custom_branches):
    key = str(row['Bank Key']).strip().upper()
    iban = str(row['Iban']).strip().upper()
    if key in BANKS_WITH_DYNAMIC_BRANCHES:
        try:
            if len(iban) >= 11:
                branch_code = iban[8:11]
                computed_bic = f"{key}IQBA{branch_code}"
                return computed_bic
            return BANK_BICS.get(key, 'UnknownBIC')
        except:
            return BANK_BICS.get(key, 'UnknownBIC')
    return BANK_BICS.get(key, 'UnknownBIC')


def adjust_column_width_ws(ws):
    no_border = Border()
    for i, col in enumerate(ws.columns):
        max_length = 0
        column = get_column_letter(i + 1)
        for cell in col:
            try:
                cell.border = no_border
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 3, 15)

# ====================================================================
# منطق المعالجة الرئيسي — يعيد dict من اسم الملف → bytes
# ====================================================================

def process_excel_data(df_input, col_name, col_iban, col_salary, custom_branches):
    df = df_input.copy()

    missing = [c for c in [col_name, col_iban, col_salary] if c not in df.columns]
    if missing:
        raise ValueError(f"الأعمدة التالية غير موجودة: {', '.join(missing)}")

    df = df.rename(columns={
        col_name:   'الاسم',
        col_iban:   'Iban',
        col_salary: 'الراتب الصافي'
    })

    df['الراتب الصافي'] = pd.to_numeric(df['الراتب الصافي'], errors='coerce').fillna(0)
    df = df[df['الراتب الصافي'] != 0].dropna(subset=['Iban', 'الاسم'])

    original_total = round(df['الراتب الصافي'].sum(), 2)
    decimal_count  = int((df['الراتب الصافي'] % 1 != 0).sum())
    floor_total    = df['الراتب الصافي'].apply(math.floor).sum()
    dinars_to_dist = int(round(original_total - floor_total))

    df = apply_financial_rounding(df, 'الراتب الصافي')
    final_total   = df['الراتب الصافي'].sum()
    remaining_dec = int((df['الراتب الصافي'] % 1 != 0).sum())

    df['الاسم'] = df['الاسم'].astype(str).str.split().str.join(' ').str[:35]
    today    = datetime.now()
    date_str = st.session_state.get('value_date_override') or today.strftime('%Y%m%d')
    try:
        entered_month = int(date_str[4:6])
        entered_year  = int(date_str[0:4])
    except (ValueError, IndexError):
        entered_month = today.month
        entered_year  = today.year
    month_ar = ARABIC_MONTHS.get(entered_month, "")

    df['Value Date']             = date_str
    df['Payer Name']             = PAYER_NAME
    df['Payer Account']          = PAYER_ACCOUNT
    df['Amount']                 = df['الراتب الصافي']
    df['Currency']               = CURRENCY
    df['Details of Charges']     = DETAILS_OF_CHARGES
    df['Beneficiary Name']       = df['الاسم']
    df['Beneficiary Account']    = df['Iban']
    df['Remittance Information'] = REMITTANCE_INFO_TEMPLATE.format(entered_year, month_ar)
    df['Bank Key']               = df['Iban'].str[4:8]

    df_filtered = df[df['Bank Key'].isin(BANK_KEYS_FOR_FILTERING)].copy()
    df_filtered['Receiver BIC'] = df_filtered.apply(
        lambda row: get_receiver_bic_dynamic(row, custom_branches), axis=1
    )
    df_filtered['Reference'] = date_str + ' ' + df_filtered['Iban'].astype(str)

    output_files = {}
    grouped = df_filtered.groupby('Bank Key')

    for bank_key, bank_df in grouped:
        bank_ar   = ARABIC_BANK_NAME_MAP.get(bank_key, 'مصرف')
        num_rows  = len(bank_df)
        start_row = 0
        file_index = 1

        while start_row < num_rows:
            end_row       = min(start_row + MAX_ROWS_PER_FILE, num_rows)
            current_slice = bank_df.iloc[start_row:end_row]

            while current_slice['Amount'].sum() > MAX_AMOUNT_PER_FILE and len(current_slice) > 1:
                end_row      -= 1
                current_slice = bank_df.iloc[start_row:end_row]

            filename = f"{bank_ar}_الملف_{file_index}_{bank_key}_{date_str}.xlsx"
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                current_slice[FINAL_EXCEL_COLS].to_excel(writer, index=False, sheet_name='Sheet1')
                ws = writer.sheets['Sheet1']
                adjust_column_width_ws(ws)
            buf.seek(0)
            output_files[filename] = buf.read()

            start_row   = end_row
            file_index += 1

    stats = {
        'original_total':  original_total,
        'final_total':     final_total,
        'decimal_count':   decimal_count,
        'dinars_to_dist':  dinars_to_dist,
        'remaining_dec':   remaining_dec,
        'file_count':      len(output_files),
    }
    return output_files, stats


# ====================================================================
# إنشاء ملف الملخص
# ====================================================================

def create_summary_file(excel_files_dict, use_arabic_digits=False):
    """
    excel_files_dict: dict من اسم الملف → bytes (ملفات xlsx المقسمة)
    يعيد bytes لملف xlsx الملخص
    """
    rows_data = []
    for filename, file_bytes in excel_files_dict.items():
        df    = pd.read_excel(io.BytesIO(file_bytes))
        parts = filename.split('_')
        bank_name   = parts[0]
        branch_code = parts[3] if len(parts) > 3 else "---"
        total_amount = int(df['Amount'].sum()) if 'Amount' in df.columns else 0
        rows_data.append({
            'المصرف':          bank_name,
            'اسم الملف':       filename,
            'رمز الفرع':       branch_code,
            'عدد الموظفين':    len(df),
            'المبلغ الإجمالي': total_amount,
        })

    rows_data.sort(key=lambda x: x['المصرف'])

    COLOR_HEADER      = "1F4E78"
    COLOR_BANK_SUBTOT = "2E75B6"
    COLOR_GRAND_TOTAL = "C00000"
    COLOR_ROW_ODD     = "EBF3FB"
    COLOR_ROW_EVEN    = "FFFFFF"
    COLOR_BANK_HEADER = "D6E4F0"

    thin_border   = Border(left=Side(style='thin'),   right=Side(style='thin'),
                           top=Side(style='thin'),    bottom=Side(style='thin'))
    medium_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                           top=Side(style='medium'),  bottom=Side(style='medium'))
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align   = Alignment(horizontal='right',  vertical='center', wrap_text=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='الملخص الشامل', index=False)
        ws = writer.sheets['الملخص الشامل']

        col_widths  = [18, 50, 12, 16, 22, 65]
        col_letters = ['A', 'B', 'C', 'D', 'E', 'F']
        for letter, width in zip(col_letters, col_widths):
            ws.column_dimensions[letter].width = width

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

        ws.merge_cells('A1:F1')
        tc = ws['A1']
        tc.value     = f"ملخص الحسابات الشامل — {PAYER_NAME}"
        tc.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
        tc.fill      = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        tc.alignment = Alignment(horizontal='center', vertical='center')
        tc.border    = medium_border

        headers = ['المصرف', 'اسم الملف المرجعي', 'رمز الفرع', 'عدد الموظفين', 'المبلغ الإجمالي (د.ع)', 'المبلغ كتابةً']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
            cell.fill      = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.alignment = center_align
            cell.border    = thin_border

        current_row = 3
        grand_total_employees = 0
        grand_total_amount    = 0

        for bank_name, bank_rows in groupby(rows_data, key=lambda x: x['المصرف']):
            bank_rows_list = list(bank_rows)
            bank_employees = sum(r['عدد الموظفين']    for r in bank_rows_list)
            bank_amount    = sum(r['المبلغ الإجمالي'] for r in bank_rows_list)

            ws.row_dimensions[current_row].height = 22
            ws.merge_cells(f'A{current_row}:F{current_row}')
            btc = ws[f'A{current_row}']
            btc.value     = f"🏦  مصرف  {bank_name}"
            btc.font      = Font(bold=True, size=12, color="1F4E78", name="Arial")
            btc.fill      = PatternFill(start_color=COLOR_BANK_HEADER, end_color=COLOR_BANK_HEADER, fill_type="solid")
            btc.alignment = Alignment(horizontal='right', vertical='center')
            btc.border    = medium_border
            current_row  += 1

            for file_idx, row_data in enumerate(bank_rows_list):
                ws.row_dimensions[current_row].height = 18
                bg_color = COLOR_ROW_ODD if file_idx % 2 == 0 else COLOR_ROW_EVEN
                values = [
                    row_data['المصرف'],
                    row_data['اسم الملف'],
                    row_data['رمز الفرع'],
                    row_data['عدد الموظفين'],
                    row_data['المبلغ الإجمالي'],
                    convert_amount_to_arabic(row_data['المبلغ الإجمالي']),
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.fill      = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    cell.border    = thin_border
                    cell.font      = Font(size=10, name="Arial")
                    cell.alignment = center_align if col_idx != 6 else right_align
                    if col_idx == 5:
                        if use_arabic_digits:
                            cell.value     = format_number(val, use_arabic_digits=True)
                            cell.alignment = center_align
                        else:
                            cell.number_format = '#,##0'
                current_row += 1

            ws.row_dimensions[current_row].height = 24
            subtot_labels = [
                f"إجمالي {bank_name}", "───", "───",
                bank_employees, bank_amount,
                convert_amount_to_arabic(bank_amount),
            ]
            for col_idx, val in enumerate(subtot_labels, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
                cell.fill      = PatternFill(start_color=COLOR_BANK_SUBTOT, end_color=COLOR_BANK_SUBTOT, fill_type="solid")
                cell.border    = medium_border
                cell.alignment = center_align if col_idx != 6 else right_align
                if col_idx == 5:
                    if use_arabic_digits:
                        cell.value = format_number(val, use_arabic_digits=True)
                    else:
                        cell.number_format = '#,##0'
            current_row += 1

            ws.row_dimensions[current_row].height = 8
            for col_idx in range(1, 7):
                ws.cell(row=current_row, column=col_idx).fill = PatternFill(
                    start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
                )
            current_row += 1

            grand_total_employees += bank_employees
            grand_total_amount    += bank_amount

        ws.row_dimensions[current_row].height = 30
        grand_labels = [
            "✦ المجموع الكلي النهائي ✦", "───", "───",
            grand_total_employees, grand_total_amount,
            convert_amount_to_arabic(grand_total_amount),
        ]
        for col_idx, val in enumerate(grand_labels, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font      = Font(bold=True, size=12, color="FFFFFF", name="Arial")
            cell.fill      = PatternFill(start_color=COLOR_GRAND_TOTAL, end_color=COLOR_GRAND_TOTAL, fill_type="solid")
            cell.border    = medium_border
            cell.alignment = center_align if col_idx != 6 else right_align
            if col_idx == 5:
                if use_arabic_digits:
                    cell.value = format_number(val, use_arabic_digits=True)
                else:
                    cell.number_format = '#,##0'
        current_row += 1

        ws.row_dimensions[current_row].height = 28
        ws.merge_cells(f'A{current_row}:F{current_row}')
        wc = ws[f'A{current_row}']
        wc.value     = f"المبلغ الكلي كتابةً :  {convert_amount_to_arabic(grand_total_amount)}"
        wc.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        wc.fill      = PatternFill(start_color="7B2C2C", end_color="7B2C2C", fill_type="solid")
        wc.alignment = Alignment(horizontal='right', vertical='center')
        wc.border    = medium_border

        ws.freeze_panes         = 'A3'
        ws.sheet_view.rightToLeft = True

    buf.seek(0)
    return buf.read(), grand_total_employees, grand_total_amount


# ====================================================================
# تحويل Excel → CSV بنكي
# ====================================================================

def convert_excel_to_csv_bytes(file_bytes):
    df  = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    out = io.BytesIO()
    header = '|'.join(FINAL_EXCEL_COLS) + '\r\n'
    out.write(header.encode('utf-8'))
    for _, row in df[FINAL_EXCEL_COLS].iterrows():
        line_values = []
        for col in FINAL_EXCEL_COLS:
            raw_val = str(row[col]).strip()
            if col == 'Amount':
                try:
                    clean = raw_val.replace(',', '').replace('"', '')
                    val   = str(int(float(clean)))
                except:
                    val = '0'
            elif col == 'Beneficiary Name':
                clean_name = ' '.join(raw_val.split())
                val        = " " + clean_name[:35]
            else:
                val = raw_val
            line_values.append(val)
        out.write(('|'.join(line_values) + '\r\n').encode('utf-8'))
    out.seek(0)
    return out.read()


# ====================================================================
# واجهة Streamlit
# ====================================================================

st.set_page_config(
    page_title="نظام رواتب تربية البصرة",
    page_icon="🏛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS للدعم الكامل للعربية وRTL
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');

    html, body, [class*="css"], .stApp {
        font-family: 'Cairo', 'Arial', sans-serif !important;
        direction: rtl;
    }

    /* Header */
    .main-header {
        background: linear-gradient(135deg, #1A237E 0%, #283593 100%);
        padding: 28px 24px;
        border-radius: 12px;
        margin-bottom: 24px;
        text-align: center;
        border: 1px solid #3949AB;
    }
    .main-header h1 {
        color: #FFD700;
        font-size: 26px;
        font-weight: 700;
        margin: 0 0 6px 0;
    }
    .main-header p {
        color: #90CAF9;
        font-size: 14px;
        margin: 0;
    }

    /* Section headers */
    .section-head {
        background: linear-gradient(90deg, #1F3A5F, #243B55);
        color: #90CAF9;
        font-weight: 700;
        font-size: 15px;
        padding: 10px 18px;
        border-radius: 8px;
        margin: 18px 0 10px 0;
        border-right: 4px solid #FFD700;
    }

    /* Stats cards */
    .stat-card {
        background: #1E2A3A;
        border: 1px solid #2E4057;
        border-radius: 10px;
        padding: 16px;
        text-align: center;
        margin: 4px;
    }
    .stat-card .label {
        color: #90CAF9;
        font-size: 12px;
        margin-bottom: 6px;
    }
    .stat-card .value {
        color: #FFD700;
        font-size: 20px;
        font-weight: 700;
    }

    /* Success box */
    .success-box {
        background: #1B3A2D;
        border: 1px solid #2ECC71;
        border-radius: 10px;
        padding: 16px 20px;
        margin: 12px 0;
        color: #A8EDCA;
    }

    /* Branch tag */
    .branch-tag {
        display: inline-block;
        background: #1F3A5F;
        color: #90CAF9;
        border: 1px solid #2E75B6;
        border-radius: 6px;
        padding: 4px 12px;
        font-size: 13px;
        margin: 3px;
        font-family: monospace;
    }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: #111827;
        border-left: 1px solid #1F3A5F;
    }
    [data-testid="stSidebar"] .stMarkdown, 
    [data-testid="stSidebar"] label {
        color: #CBD5E1 !important;
        direction: rtl;
    }

    /* Inputs */
    .stSelectbox label, .stFileUploader label, .stRadio label {
        color: #CBD5E1 !important;
        font-weight: 600;
    }

    /* Buttons */
    .stButton > button {
        background: #0D47A1;
        color: white;
        border: none;
        border-radius: 8px;
        font-family: 'Cairo', sans-serif;
        font-weight: 600;
        font-size: 14px;
        padding: 10px 20px;
        width: 100%;
        transition: all 0.2s;
    }
    .stButton > button:hover {
        background: #1565C0;
        transform: translateY(-1px);
    }

    /* Download button */
    .stDownloadButton > button {
        background: #1B5E20;
        color: white;
        border: none;
        border-radius: 8px;
        font-family: 'Cairo', sans-serif;
        font-weight: 600;
        width: 100%;
        padding: 10px 20px;
    }
    .stDownloadButton > button:hover {
        background: #2E7D32;
    }

    /* Hide Streamlit branding */
    #MainMenu, footer { visibility: hidden; }
    .block-container { padding-top: 1rem; }

    /* Tables */
    .stDataFrame { direction: rtl; }

    div[data-testid="stMetricValue"] {
        color: #FFD700 !important;
        font-family: 'Cairo', sans-serif !important;
    }
</style>
""", unsafe_allow_html=True)

# ====================================================================
# تهيئة Session State
# ====================================================================

if 'custom_branches' not in st.session_state:
    st.session_state.custom_branches = DEFAULT_BRANCHES.copy()

if 'processed_files' not in st.session_state:
    st.session_state.processed_files = {}

if 'process_stats' not in st.session_state:
    st.session_state.process_stats = None

if 'df_preview' not in st.session_state:
    st.session_state.df_preview = None

if 'df_columns' not in st.session_state:
    st.session_state.df_columns = []

if 'value_date_override' not in st.session_state:
    st.session_state.value_date_override = None

# ====================================================================
# Header
# ====================================================================

st.markdown("""
<div class="main-header">
    <h1>🏛 نظام معالجة رواتب تربية البصرة</h1>
    <p>جبر مالي  •  تقسيم مصرفي  •  ملخص منسق  •  تشفير CSV</p>
</div>
""", unsafe_allow_html=True)

# ====================================================================
# Sidebar — الإعدادات والفروع
# ====================================================================

with st.sidebar:
    st.markdown("## ⚙️ الإعدادات")
    st.markdown("---")

    st.markdown("### 🏦 إدارة فروع المصارف")
    st.caption("(الفروع الديناميكية: AINI, NBIQ, AIBI)")

    sel_bank = st.selectbox("اختر المصرف", BANKS_WITH_DYNAMIC_BRANCHES, key="sb_bank")

    current_bics = st.session_state.custom_branches.get(sel_bank, [])
    if current_bics:
        st.markdown("**الفروع الحالية:**")
        for bic in current_bics:
            st.markdown(f'<span class="branch-tag">📌 {bic}</span>', unsafe_allow_html=True)
    else:
        st.caption("لا توجد فروع مضافة")

    new_bic = st.text_input("BIC الفرع الجديد", placeholder="مثال: NBIQIQBA863", key="new_bic")
    if st.button("➕ إضافة فرع"):
        bic_clean = new_bic.strip().upper()
        if len(bic_clean) < 8:
            st.error("BIC يجب أن يكون 8 أحرف على الأقل")
        else:
            if sel_bank not in st.session_state.custom_branches:
                st.session_state.custom_branches[sel_bank] = []
            if bic_clean in st.session_state.custom_branches[sel_bank]:
                st.warning("هذا الفرع موجود مسبقاً")
            else:
                st.session_state.custom_branches[sel_bank].append(bic_clean)
                st.success(f"✔ تمت إضافة {bic_clean}")
                st.rerun()

    del_bic = st.text_input("BIC للحذف", placeholder="مثال: NBIQIQBA830", key="del_bic")
    if st.button("🗑 حذف فرع"):
        bic_del = del_bic.strip().upper()
        if sel_bank in st.session_state.custom_branches and bic_del in st.session_state.custom_branches[sel_bank]:
            st.session_state.custom_branches[sel_bank].remove(bic_del)
            st.success(f"✔ تم حذف {bic_del}")
            st.rerun()
        else:
            st.error("الفرع غير موجود")

    st.markdown("---")
    st.markdown("### 📊 خيارات الملخص")
    digit_mode = st.radio(
        "نوع الأرقام في الملخص",
        ["أرقام إنجليزية  1,000", "أرقام عربية  ١،٠٠٠"],
        key="digit_mode"
    )
    use_arabic_digits = ("عربية" in digit_mode)

    st.markdown("---")
    st.markdown("### 📅 تاريخ التحويل")
    custom_date = st.date_input("تاريخ الإرسال للبنك", value=datetime.now())
    st.session_state.value_date_override = custom_date.strftime('%Y%m%d')

    st.markdown("---")
    st.markdown("### ℹ️ معلومات النظام")
    st.caption(f"**الجهة الدافعة:** {PAYER_NAME}")
    st.caption(f"**الحساب:** {PAYER_ACCOUNT[:15]}...")
    st.caption(f"**الحد الأقصى للملف:** {MAX_ROWS_PER_FILE:,} صف")

# ====================================================================
# المنطقة الرئيسية — تبويبات
# ====================================================================

tab1, tab2, tab3 = st.tabs(["① رفع الملف والمعالجة", "② الملخص", "③ التشفير CSV"])

# ─────────────────────────────────────────────
# التبويب الأول: رفع ومعالجة
# ─────────────────────────────────────────────
with tab1:
    st.markdown('<div class="section-head">① رفع ملف Excel</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "اختر ملف Excel (.xlsx)",
        type=["xlsx"],
        key="uploader"
    )

    if uploaded:
        try:
            df_raw = pd.read_excel(uploaded)
            st.session_state.df_preview = df_raw
            st.session_state.df_columns = list(df_raw.columns)
            st.success(f"✔ تم تحميل الملف | {len(df_raw):,} صف | {len(df_raw.columns)} عمود")

            with st.expander("👁 معاينة أول 5 صفوف"):
                st.dataframe(df_raw.head(), use_container_width=True)
        except Exception as e:
            st.error(f"خطأ في قراءة الملف: {e}")

    if st.session_state.df_columns:
        st.markdown('<div class="section-head">② تحديد الأعمدة</div>', unsafe_allow_html=True)

        cols_list = st.session_state.df_columns

        default_name   = guess_column(['اسم', 'name', 'الاسم', 'موظف'], cols_list)
        default_iban   = guess_column(['iban', 'حساب', 'account', 'رقم'], cols_list)
        default_salary = guess_column(['راتب', 'salary', 'صافي', 'net', 'مبلغ', 'amount'], cols_list)

        c1, c2, c3 = st.columns(3)
        with c1:
            col_name = st.selectbox("👤 عمود الاسم", cols_list,
                                    index=cols_list.index(default_name) if default_name in cols_list else 0)
        with c2:
            col_iban = st.selectbox("🏦 عمود IBAN", cols_list,
                                    index=cols_list.index(default_iban) if default_iban in cols_list else 0)
        with c3:
            col_salary = st.selectbox("💰 عمود الراتب الصافي", cols_list,
                                      index=cols_list.index(default_salary) if default_salary in cols_list else 0)

        if len({col_name, col_iban, col_salary}) < 3:
            st.warning("⚠ لا يمكن اختيار نفس العمود لأكثر من حقل!")
        else:
            st.markdown('<div class="section-head">③ بدء المعالجة</div>', unsafe_allow_html=True)

            if st.button("🚀 بدء التقسيم والمعالجة (جبر مالي)", use_container_width=True):
                with st.spinner("جاري المعالجة والجبر المالي..."):
                    try:
                        files, stats = process_excel_data(
                            st.session_state.df_preview,
                            col_name, col_iban, col_salary,
                            st.session_state.custom_branches
                        )
                        st.session_state.processed_files = files
                        st.session_state.process_stats   = stats
                    except Exception as e:
                        st.error(f"خطأ: {e}")

    if st.session_state.process_stats:
        stats = st.session_state.process_stats
        st.markdown("---")
        st.markdown("### ✅ نتائج المعالجة")

        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("الملفات المنشأة",     str(stats['file_count']))
        m2.metric("الصفوف ذات أعشار",   f"{stats['decimal_count']:,}")
        m3.metric("الدنانير الموزعة",    f"{stats['dinars_to_dist']:,}")
        m4.metric("المجموع الأصلي",      f"{stats['original_total']:,.2f}")
        m5.metric("المجموع بعد الجبر",   f"{stats['final_total']:,}")

        diff = stats['final_total'] - stats['original_total']
        rc = "أعشار متبقية"
        st.markdown(f"""
        <div class="success-box">
        ✔ الجبر المالي ناجح | الفرق: <b>{diff:+.2f}</b> | {rc}: <b>{stats['remaining_dec']}</b>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("### 📥 تحميل الملفات المقسمة")

        if len(st.session_state.processed_files) == 1:
            fname, fbytes = list(st.session_state.processed_files.items())[0]
            st.download_button(f"⬇ تحميل {fname}", fbytes, fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, fbytes in st.session_state.processed_files.items():
                    zf.writestr(fname, fbytes)
            zip_buf.seek(0)
            date_str = datetime.now().strftime('%Y%m%d')
            st.download_button(
                f"⬇ تحميل جميع الملفات ({len(st.session_state.processed_files)} ملف) — ZIP",
                zip_buf.read(),
                f"payroll_files_{date_str}.zip",
                mime="application/zip",
                use_container_width=True
            )
            st.markdown("**أو تحميل كل ملف منفرداً:**")
            for fname, fbytes in st.session_state.processed_files.items():
                st.download_button(
                    f"⬇ {fname}", fbytes, fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{fname}"
                )

# ─────────────────────────────────────────────
# التبويب الثاني: الملخص
# ─────────────────────────────────────────────
with tab2:
    st.markdown('<div class="section-head">إنشاء ملف الملخص المفصّل</div>', unsafe_allow_html=True)

    if not st.session_state.processed_files:
        st.info("ℹ قم بمعالجة الملف في التبويب الأول أولاً، ثم عد هنا لإنشاء الملخص.")
    else:
        files_in  = st.session_state.processed_files
        n_files   = len(files_in)
        st.write(f"**الملفات المتاحة للتلخيص:** {n_files} ملف")

        # جدول معاينة سريعة
        preview_rows = []
        for fname, fbytes in files_in.items():
            df_tmp = pd.read_excel(io.BytesIO(fbytes))
            parts  = fname.split('_')
            preview_rows.append({
                "المصرف":         parts[0],
                "اسم الملف":      fname,
                "عدد الموظفين":   len(df_tmp),
                "المبلغ الإجمالي": int(df_tmp['Amount'].sum()) if 'Amount' in df_tmp.columns else 0,
            })
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

        if st.button("📊 إنشاء ملف الملخص", use_container_width=True):
            with st.spinner("جاري إنشاء الملخص..."):
                try:
                    summary_bytes, total_emp, total_amt = create_summary_file(
                        files_in, use_arabic_digits
                    )
                    fname_summary = f"Summary_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                    st.success(f"✔ تم إنشاء الملخص | {total_emp:,} موظف | {total_amt:,} د.ع")
                    st.markdown(f"**المبلغ كتابةً:** {convert_amount_to_arabic(total_amt)}")

                    st.download_button(
                        "⬇ تحميل ملف الملخص",
                        summary_bytes,
                        fname_summary,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"خطأ في إنشاء الملخص: {e}")

# ─────────────────────────────────────────────
# التبويب الثالث: التشفير CSV
# ─────────────────────────────────────────────
with tab3:
    st.markdown('<div class="section-head">تشفير وتحويل الملفات إلى CSV البنكي</div>', unsafe_allow_html=True)

    if not st.session_state.processed_files:
        st.info("ℹ قم بمعالجة الملف في التبويب الأول أولاً.")
    else:
        st.write(f"**الملفات الجاهزة للتشفير:** {len(st.session_state.processed_files)} ملف")
        st.caption("يتم تحويل كل ملف Excel إلى CSV بصيغة بنكية مطابقة (UTF-8, CRLF, مفصولة بـ |)")

        if st.button("🔑 تشفير وتحويل جميع الملفات → CSV", use_container_width=True):
            with st.spinner("جاري التشفير..."):
                try:
                    csv_files = {}
                    for fname, fbytes in st.session_state.processed_files.items():
                        csv_bytes = convert_excel_to_csv_bytes(fbytes)
                        csv_name  = os.path.splitext(fname)[0] + ".csv"
                        csv_files[csv_name] = csv_bytes

                    zip_buf = io.BytesIO()
                    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for cname, cbytes in csv_files.items():
                            zf.writestr(cname, cbytes)
                    zip_buf.seek(0)

                    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
                    st.success(f"✔ تم تشفير {len(csv_files)} ملف بنجاح")

                    st.download_button(
                        f"⬇ تحميل ملفات CSV ({len(csv_files)} ملف) — ZIP",
                        zip_buf.read(),
                        f"csv_encrypted_{date_str}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )

                    st.markdown("**أو تحميل كل ملف CSV منفرداً:**")
                    for cname, cbytes in csv_files.items():
                        st.download_button(
                            f"⬇ {cname}", cbytes, cname,
                            mime="text/csv",
                            key=f"csv_{cname}"
                        )
                except Exception as e:
                    st.error(f"خطأ في التشفير: {e}")
